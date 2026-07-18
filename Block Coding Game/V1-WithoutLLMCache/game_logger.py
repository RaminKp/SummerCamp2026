"""
Game logging — writes three formats in parallel:
  game_log.json  — flat array of session objects (one per game)
  game_log.xlsx  — Excel workbook (Sessions / Attempts / PlayerSummary sheets)
  logs/sessions.csv       — one row per game
  logs/attempts.csv       — one row per RFID submission
  logs/player_summary.csv — one row per game, rolled-up accuracy
"""

import csv
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

BASE_DIR  = Path(__file__).parent
LOG_PATH  = BASE_DIR / "game_log.xlsx"
JSON_PATH = BASE_DIR / "game_log.json"
CSV_DIR   = BASE_DIR / "logs"

SESSIONS_HEADER = [
    "SessionID",
    "Player1 Name", "Player1 ArUco", "Player1 Age", "Player1 Plays",
    "Player2 Name", "Player2 ArUco", "Player2 Age", "Player2 Plays",
    "Date", "Map", "Start Time", "End Time", "Duration (s)", "Outcome",
]

ATTEMPTS_HEADER = [
    "SessionID", "Player1 Name", "Player2 Name",
    "Date", "Map", "Checkpoint", "Attempt #",
    "Scanned RFIDs", "Expected RFIDs", "Result", "Correct?",
    "Start Time", "End Time", "Duration (s)",
]

SUMMARY_HEADER = [
    "SessionID",
    "Player1 Name", "Player1 ArUco",
    "Player2 Name", "Player2 ArUco",
    "Date", "Map", "Outcome",
    "Total Attempts", "Correct Attempts", "Incorrect Attempts",
    "Accuracy (%)", "Game Duration (s)",
]


# ── JSON helpers ──────────────────────────────────────────────────────────────

def _load_json() -> list:
    if JSON_PATH.exists():
        try:
            data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
            if isinstance(data, list):
                return data
        except Exception:
            pass
    return []


def _save_json(sessions: list):
    JSON_PATH.write_text(json.dumps(sessions, indent=2, ensure_ascii=False),
                         encoding="utf-8")


# ── CSV helpers ───────────────────────────────────────────────────────────────

def _append_csv(filename: str, header: list, row: list):
    CSV_DIR.mkdir(exist_ok=True)
    path   = CSV_DIR / filename
    is_new = not path.exists()
    with open(path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if is_new:
            writer.writerow(header)
        writer.writerow(row)


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _ensure_workbook() -> Workbook:
    if LOG_PATH.exists():
        wb = load_workbook(LOG_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "Sessions" not in wb.sheetnames:
        wb.create_sheet("Sessions").append(SESSIONS_HEADER)
    if "Attempts" not in wb.sheetnames:
        wb.create_sheet("Attempts").append(ATTEMPTS_HEADER)
    if "PlayerSummary" not in wb.sheetnames:
        wb.create_sheet("PlayerSummary").append(SUMMARY_HEADER)

    return wb


class GameLogger:
    """One instance per game. Call start(), log_attempt()*, end()."""

    def __init__(self, players: list[dict], map_name: str):
        """
        players: list of 2 dicts from id_scanner.wait_for_players()
                 each has keys: aruco_id, name, age, plays
        """
        self.players   = players
        self.map_name  = map_name
        self.session_id: str | None = None
        self._session_start: datetime | None = None
        self._checkpoint_start: datetime | None = None
        self._checkpoints: list[dict] = []   # inline in JSON session object
        self._attempt_rows: list[dict] = []  # for Excel / CSV / accuracy

    def _p(self, idx: int) -> dict:
        return self.players[idx] if idx < len(self.players) else {}

    # ── session-level ──────────────────────────────────────────────────────

    def start(self):
        self._session_start = datetime.now()
        p1 = self._p(0)
        p2 = self._p(1)
        self.session_id = (
            f"{self._session_start.strftime('%Y-%m-%dT%H:%M:%S')}"
            f"_{p1.get('aruco_id', 'X')}_{p2.get('aruco_id', 'X')}"
        )
        print(f"  Session ID: {self.session_id}")

    def end(self, outcome: str):
        end_time = datetime.now()
        start    = self._session_start or end_time
        duration = round((end_time - start).total_seconds())
        p1, p2   = self._p(0), self._p(1)

        total    = len(self._attempt_rows)
        correct  = sum(1 for a in self._attempt_rows if a["correct"])
        accuracy = round(100 * correct / total, 1) if total else 0.0

        # ── JSON — flat array, one session object with inline checkpoints ──
        session_obj = {
            "session_id": self.session_id,
            "team": {
                "player_1": {
                    "aruco_id":      str(p1.get("aruco_id", "")),
                    "name":          p1.get("name", ""),
                    "age":           p1.get("age", ""),
                    "attempt_number": p1.get("plays", 0),
                },
                "player_2": {
                    "aruco_id":      str(p2.get("aruco_id", "")),
                    "name":          p2.get("name", ""),
                    "age":           p2.get("age", ""),
                    "attempt_number": p2.get("plays", 0),
                },
            },
            "map":                   self.map_name,
            "game_start":            start.strftime("%Y-%m-%dT%H:%M:%S"),
            "game_end":              end_time.strftime("%Y-%m-%dT%H:%M:%S"),
            "total_duration_seconds": duration,
            "checkpoints":           self._checkpoints,
            "outcome":               outcome,
            "correct_attempts":      correct,
            "total_score":           float(correct),
        }
        sessions = _load_json()
        sessions.append(session_obj)
        _save_json(sessions)

        # ── Excel ──────────────────────────────────────────────────────────
        session_row = [
            self.session_id,
            p1.get("name", ""),  p1.get("aruco_id", ""), p1.get("age", ""), p1.get("plays", 0),
            p2.get("name", ""),  p2.get("aruco_id", ""), p2.get("age", ""), p2.get("plays", 0),
            start.strftime("%Y-%m-%d"),
            self.map_name,
            start.strftime("%H:%M:%S"),
            end_time.strftime("%H:%M:%S"),
            duration,
            outcome,
        ]
        summary_row = [
            self.session_id,
            p1.get("name", ""), p1.get("aruco_id", ""),
            p2.get("name", ""), p2.get("aruco_id", ""),
            start.strftime("%Y-%m-%d"),
            self.map_name,
            outcome,
            total, correct, total - correct, accuracy, duration,
        ]
        wb = _ensure_workbook()
        wb["Sessions"].append(session_row)
        wb["PlayerSummary"].append(summary_row)
        wb.save(LOG_PATH)

        # ── CSV ────────────────────────────────────────────────────────────
        _append_csv("sessions.csv",       SESSIONS_HEADER, session_row)
        _append_csv("player_summary.csv", SUMMARY_HEADER,  summary_row)

    # ── per-checkpoint-attempt level ───────────────────────────────────────

    def begin_checkpoint_attempt(self):
        self._checkpoint_start = datetime.now()

    def log_attempt(self, checkpoint_label: str, attempt_num: int,
                    scanned: list[int], expected: list[int], result: str):
        end_time = datetime.now()
        start    = self._checkpoint_start or end_time
        duration = round((end_time - start).total_seconds())
        correct  = (result == "CORRECT")

        self._attempt_rows.append({"correct": correct})

        # ── JSON — checkpoint entry stored inline on session object ────────
        self._checkpoints.append({
            "checkpoint":       checkpoint_label,
            "attempt":          attempt_num,
            "start_time":       start.strftime("%Y-%m-%dT%H:%M:%S"),
            "end_time":         end_time.strftime("%Y-%m-%dT%H:%M:%S"),
            "duration_seconds": duration,
            "result":           result.lower(),
            "scanned_rfids":    scanned,
            "expected_rfids":   expected,
        })

        p1, p2 = self._p(0), self._p(1)
        attempt_row = [
            self.session_id,
            p1.get("name", ""), p2.get("name", ""),
            start.strftime("%Y-%m-%d"),
            self.map_name,
            checkpoint_label,
            attempt_num,
            str(scanned),
            str(expected),
            result,
            "Yes" if correct else "No",
            start.strftime("%H:%M:%S"),
            end_time.strftime("%H:%M:%S"),
            duration,
        ]

        # ── Excel ──────────────────────────────────────────────────────────
        wb = _ensure_workbook()
        wb["Attempts"].append(attempt_row)
        wb.save(LOG_PATH)

        # ── CSV ────────────────────────────────────────────────────────────
        _append_csv("attempts.csv", ATTEMPTS_HEADER, attempt_row)
